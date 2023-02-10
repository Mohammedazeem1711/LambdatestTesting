from time import sleep
import openpyxl
from selenium import webdriver
from Locators_keys import Commonkeys
from Locator_values import Login_Locators, preference, Reports, usradmin
from Selenium.Jumpcharge.Path import URLPath
from Selenium.Jumpcharge.URL import JumpchargeUrl



class BaseClassMethods:

    def __init__(self):
        self.driver = webdriver.Chrome(URLPath.pathchr)

    def openQAUrl(self):
        self.driver.get(JumpchargeUrl.produrl)
        self.driver.maximize_window()

    def readData(self, file, sheetname, rownum, colnum):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetname]
        inputvalue = sheet.cell(row=rownum, column=colnum).value
        return inputvalue

    def openSheet(self, sheetname, emailrownum):
        sheetpath = "C:\\Users\\cprakash\\OneDrive - HUBBELL INC\\Documents\\cp\\Personal\\Others\\PycharmProjects\\Selenium\\Test_Data.xlsx"
        email = self.readData(sheetpath, sheetname, emailrownum, 1)
        pwd = self.readData(sheetpath, sheetname, emailrownum, 2)
        return email, pwd
        # email = ws.cell(2, 1).value
        # email = ws.cell(row=rownum).value
        # pwd = ws.cell(column=col).value

    def with_(self, param, param1):
        lockey = Commonkeys(self.driver)
        lockey.findByXpath(Login_Locators.email_xpath).send_keys(param)
        lockey.findByXpath(Login_Locators.password_xpath).send_keys(param1)
        element = lockey.findByXpath(Login_Locators.login_xpath)
        self.driver.execute_script("arguments[0].click()", element)


    def mydock_(self):
        lockey = Commonkeys(self.driver)
        sleep(3)
        doc = lockey.findByXpath(Login_Locators.dock_xpath)
        self.driver.execute_script("arguments[0].click()", doc)
        sleep(3)
        apd = lockey.findByXpath(Login_Locators.apdtab)
        self.driver.execute_script("arguments[0].click()", apd)
        lockey.findByXpath(Login_Locators.cotab).click()
        lockey.findByXpath(Login_Locators.bpurtab).click()
        lockey.findByXpath(Login_Locators.loctab).click()
        lockey.findByXpath(Login_Locators.homebr).click()
        sleep(3)


    def preference(self):
        prefkey = Commonkeys(self.driver)
        prefkey.findByXpath(preference.Prefmenu).click()
        sleep(3)
        prefkey.findByXpath(preference.freetab).click()
        prefkey.findByXpath(preference.usertab).click()
        prefkey.findByXpath(Login_Locators.homebr).click()
        sleep(3)

    '''def reports(self):
        rep = Commonkeys(self.driver)
        rep.findByXpath(Reports.repmenu).click()
        rep.findByXpath(Login_Locators.homebr).click()
        sleep(3)'''

    def usrmgmt(self):
        usrmgmt = Commonkeys(self.driver)
        usrmgmt.findByXpath(usradmin.usrmenu).click()
        sleep(3)
        usrmgmt.findByXpath(usradmin.addusr).click()
        usrmgmt.findByXpath(usradmin.emailusr).send_keys("test@test.com")
        usrmgmt.findByXpath(usradmin.savbtnusr).click()
        usrmgmt.findByXpath(Login_Locators.logout).click()




